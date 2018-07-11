# Copyright (c) 2018 William Lees

# Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated
# documentation files (the "Software"), to deal in the Software without restriction, including without limitation the
# rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to permit
# persons to whom the Software is furnished to do so, subject to the following conditions:

# The above copyright notice and this permission notice shall be included in all copies or substantial portions of the
# Software.

# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE
# WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR
# COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR
# OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.

# Create an Excel-based Submission Sheet from the yaml schema

import sys
import argparse
import yaml
import yamlordereddictloader
from copy import copy
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, GradientFill
from openpyxl.worksheet.table import Table, TableStyleInfo


import textwrap


def main(argv):
    parser = argparse.ArgumentParser(description='Create an Excel-based Submission Sheet from a yaml schema.')
    parser.add_argument('schemafile', help='Schema file (.yaml)')
    parser.add_argument('outfile', help='Output file (.xlsx)')
    parser.add_argument('template', help='Template file (.xlsx)')
    args = parser.parse_args()

    schema = yaml.load(open(args.schemafile, 'r'), Loader=yamlordereddictloader.Loader)
    tb = load_workbook(args.template)
    wb = Workbook()

    tag = re.compile('^\{\{.*\:.*\}\}$')

    tablenum = 1
    sheetnames = []
    for ts in tb:
        rownum=1
        ws = wb.create_sheet(title=ts.title)
        sheetnames.append(ts.title)
        for tr in ts.rows:
            new_row = []
            for cell in tr:
                if cell.value is not None and tag.match(cell.value):
                    try:
                        obj, att = cell.value[2:-2].split(':')
                        exclusions = []
                        additions = []
                        plex = None
                        if '#' in att:
                            att,plex = att.split('#')
                            plex = int(plex)
                        if '+' in att:
                            att,additions = att.split('+')
                            additions = additions.split(',')
                        if '!' in att:
                            att,exclusions = att.split('!')
                            exclusions = exclusions.split(',')
                        if obj not in schema:
                            print('Error: object %s not in schema.' % obj)
                            quit()
                        if att not in schema[obj]:
                            print('Error: attribute %s not declared in object %s.' % (att, obj))
                            quit()
                        if 'properties' not in att:
                            paste_cell(ws, rownum, cell.col_idx, schema[obj][att], cell)
                        elif plex is None:
                            first_row =  rownum
                            first_col = cell.column
                            row_headers = ['Field']
                            for k,v in schema[obj]['properties'].items():
                                for p_k, p_v in v.items():
                                    if p_k not in row_headers and p_k not in exclusions:
                                        row_headers.append(p_k)
                            row_headers = row_headers + additions
                            for i in range(len(row_headers)):
                                nc = paste_cell(ws, rownum, cell.col_idx + i, str(row_headers[i]), cell)
                                nc.font = Font(bold=True)
                            last_col = nc.column
                            rownum += 1
                            for k,v in schema[obj]['properties'].items():
                                paste_cell(ws, rownum, cell.col_idx, k, cell)
                                for h in row_headers[1:]:
                                    if h in v:
                                        paste_cell(ws, rownum, cell.col_idx + row_headers.index(h), v[h], cell)
                                    else:
                                        paste_cell(ws, rownum, cell.col_idx + row_headers.index(h), ' ', cell)
                                rownum +=1
                            cell_range = '%s%d:%s%d' % (first_col, first_row, last_col, first_row)
                            box(ws, cell_range)
                            cell_range = '%s%d:%s%d' % (first_col, first_row, last_col, rownum-1)
                            box(ws, cell_range)
                        else:
                            first_row =  rownum
                            first_col = cell.column
                            first_col_idx = cell.col_idx
                            row_headers = []
                            for k,v in schema[obj]['properties'].items():
                                row_headers.append(k)
                            row_headers = row_headers + additions
                            for i in range(len(row_headers)):
                                nc = paste_cell(ws, rownum, cell.col_idx + i, row_headers[i], cell)
                                nc.font = Font(bold=True)
                            last_col = nc.column
                            rownum += 1
                            for i in range(plex):
                                for j in range(first_col_idx, first_col_idx + len(row_headers)):
                                    paste_cell(ws, rownum, j, ' ', cell)
                                rownum += 1
                            cell_range = '%s%d:%s%d' % (first_col, first_row, last_col, first_row)
                            box(ws, cell_range)
                            cell_range = '%s%d:%s%d' % (first_col, first_row, last_col, rownum-1)
                            box(ws, cell_range)
                    except Exception as e:
                        print('Error: unable to apply tag %s' % cell.value)
                        print(e)
                        quit()
                else:
                    paste_cell(ws, rownum, cell.col_idx, cell.value, cell)
            ws.append(new_row)
            rownum += 1

        def as_text(value):
            if value is None:
                return ""
            ret = str(value)
            if '\n' in ret:
                r = ret.split('\n')
                ret = r[0]
                for l in r:
                    if len(l) > len(ret):
                        ret = l
            return ret

        for column_cells in ws.columns:
            length = max(len(as_text(cell.value))+2 for cell in column_cells)
            if len(column_cells) > 0 and column_cells[0].column == 'D' and length < 50:
                length = 50
            ws.column_dimensions[column_cells[0].column].width = length


    for sn in wb.sheetnames:
        if sn not in sheetnames:
            del wb[sn]

    wb.save(args.outfile)

def copystyles(c1, c2):
    if c1.has_style:
        c2.font = copy(c1.font)
        c2.border = copy(c1.border)
        c2.fill = copy(c1.fill)
        c2.number_format = c1.number_format
        c2.protection = copy(c1.protection)
        c2.alignment = copy(c1.alignment)

def paste_cell(ws, row, col, value, copy_style):
    nc = ws.cell(row=row, column=col)
    if isinstance(value, str) and len(value) > 60 and '\n' not in value:
        value = textwrap.fill(value, 60)
        nc.alignment = Alignment(wrapText=True)
    nc.value = value
    copystyles(copy_style, nc)
    return nc

def box(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    style_range(ws, cell_range, border=border)


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill



if __name__=="__main__":
    main(sys.argv)
